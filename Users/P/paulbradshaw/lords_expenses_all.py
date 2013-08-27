import scraperwiki
import requests
import openpyxl
import tempfile
import time
import pprint

#BASEURL = 'http://www.parliament.uk/'
#URL = 'http://www.parliament.uk/business/lords/whos-in-the-house-of-lords/house-of-lords-expenses/'
#set a variable for the spreadsheet location
#XLS = 'http://www.parliament.uk/documents/lords-finance-office/2012-13/allowances-expenses-2012-13-month-10-january-2013.xlsx'

#This is a list of the 3 URLs we want to scrape. The line at the bottom of this scraper will loop through those.
URLS = ['http://www.parliament.uk/business/lords/whos-in-the-house-of-lords/house-of-lords-expenses/','http://www.parliament.uk/mps-lords-and-offices/members-allowances/house-of-lords/holallowances/hol-expenses04/201112/','http://www.parliament.uk/mps-lords-and-offices/members-allowances/house-of-lords/holallowances/hol-expenses04/201011/']

#create a new function called 'getworkbook' which will take 1 ingredient which it renames 'url'
def getworkbook(url):
    # Loads the xlsx file from 'url' using the 'requests' libary's 'get' and 'content' - see http://docs.python-requests.org/en/latest/user/quickstart/#make-a-request
    raw = requests.get(url, verify=False).content
    #uses the 'tempfile' library and 'NamedTemporaryFile' to store temporary information in a new variable 'f' - see http://docs.python.org/2/library/tempfile.html for more on that library
    f = tempfile.NamedTemporaryFile('wb')
    #uses '.write' to add 'raw' to that
    f.write(raw)
    #uses '.seek' to start from '0' - see http://www.tutorialspoint.com/python/file_seek.htm
    f.seek(0)
    #now use the library openpyxl's .load_workbook function to grab f.name and put into 'wb'
    wb = openpyxl.load_workbook(f.name)
    #close the temporary file
    f.close()
    #return 'wb' to whatever called this function (now go back to the function def import_data below)
    return wb

def getdata(wb, sheet='Sheet1'):
    #Turn a sheet on a workbook into an array
    data = wb.get_sheet_by_name(sheet)
    return [[unicode(cell.value).strip() for cell in row] for row in data.rows]

#This grabs the data from the URL of the spreadsheet
def import_data(linkurl):
    #create variable 'wb' containing the results of running 'getworkbook' on 'linkurl' (the spreadsheet) - see def getworkbook above
    wb = getworkbook(linkurl)
    sheet_names = wb.get_sheet_names() 
    if not sheet_names:
        return
    # assume there's only 1 sheet, and get first one
    rows = getdata(wb, sheet_names[0])
    pprint.pprint(rows)

    expenses = {}
    headings = []
    record = {}
    for num in range(0,18):
#This grabs the headings in row 6 - but the second spreadsheet has these on row 7, breaking the scraper, so the elif line grabs those
        if rows[5][0] == "Name":
            print "rows[5][num]", rows[5][num]
            #various slashes, brackets and full stops caused problems as keys, so this line replaces them all
            cleanedheading = rows[5][num].replace("/","_").replace("No.","Number").replace("(","_").replace(")","_").replace("&","and")
            print "cleaned heading:", cleanedheading 
            headings.append(cleanedheading)
            print "headings:", headings
        elif rows[6][0] == "Name":
            print "rows[6][num]", rows[6][num]
            #various slashes, brackets and full stops caused problems as keys, so this line replaces them all
            cleanedheading = rows[6][num].replace("/","_").replace("No.","Number").replace("(","_").replace(")","_").replace("&","and")
            print "cleaned heading:", cleanedheading 
            headings.append(cleanedheading)
            print "headings:", headings
        else:
            print "NEITHER ROW?", linkurl
        
    for i in range(7,805):
        print rows[i][0]
        expenses['linkurl'] = linkurl
        for num in range(0,18):
            expenses[headings[num]] = rows[i][num]
        #this caused problems because the heading 'Name' has a capital 'n' which I overlooked
        scraperwiki.sqlite.save(['Name', 'linkurl'], expenses, 'expenses')
import scraperwiki
import lxml.html

#This function finds links to all XLSX files on the page passed to it (URL)
def grabexcellinks(URL):
    #Use Scraperwiki's scrape function on 'URL', put results in new variable 'html'
    html = scraperwiki.scrape(URL)
    #Use lxml.html's fromstring function on 'html', put results in new variable 'root'
    root = lxml.html.fromstring(html)
    #use cssselect method on 'root' to grab all <a> tags within a <li> tag within <div class="rte"> - and put in a new list variable 'links'
    links = root.cssselect('div.rte li a')
    #for each item in that list variable, put it in the variable 'link'
    for link in links:
        #and print the text_content of that (after the string "link text:")
        print "link text:", link.text_content()
        #use the attrib.get method on 'link' to grab the href= attribute of the HTML, and put in new 'linkurl' variable
        linkurl = link.attrib.get('href')
        #print it
        print "link type:", linkurl[-4:]
        #if the last 4 characters of linkurl are "xlsx", then:
        if linkurl[-4:] == "xlsx":
            print "SCRAPE IT!"
            #...run the function 'import_data' on the result of adding 'http://www.parliament.uk/' to that URL (see def import_data above)
            import_data('http://www.parliament.uk/'+linkurl)

#Loop through the list 'URLS':
for URL in URLS:
    #and run the function 'grabexcellinks' on each one (see def grabexcellinks above)
    grabexcellinks(URL)

import scraperwiki
import requests
import openpyxl
import tempfile
import time
import pprint

#BASEURL = 'http://www.parliament.uk/'
#URL = 'http://www.parliament.uk/business/lords/whos-in-the-house-of-lords/house-of-lords-expenses/'
#set a variable for the spreadsheet location
#XLS = 'http://www.parliament.uk/documents/lords-finance-office/2012-13/allowances-expenses-2012-13-month-10-january-2013.xlsx'

#This is a list of the 3 URLs we want to scrape. The line at the bottom of this scraper will loop through those.
URLS = ['http://www.parliament.uk/business/lords/whos-in-the-house-of-lords/house-of-lords-expenses/','http://www.parliament.uk/mps-lords-and-offices/members-allowances/house-of-lords/holallowances/hol-expenses04/201112/','http://www.parliament.uk/mps-lords-and-offices/members-allowances/house-of-lords/holallowances/hol-expenses04/201011/']

#create a new function called 'getworkbook' which will take 1 ingredient which it renames 'url'
def getworkbook(url):
    # Loads the xlsx file from 'url' using the 'requests' libary's 'get' and 'content' - see http://docs.python-requests.org/en/latest/user/quickstart/#make-a-request
    raw = requests.get(url, verify=False).content
    #uses the 'tempfile' library and 'NamedTemporaryFile' to store temporary information in a new variable 'f' - see http://docs.python.org/2/library/tempfile.html for more on that library
    f = tempfile.NamedTemporaryFile('wb')
    #uses '.write' to add 'raw' to that
    f.write(raw)
    #uses '.seek' to start from '0' - see http://www.tutorialspoint.com/python/file_seek.htm
    f.seek(0)
    #now use the library openpyxl's .load_workbook function to grab f.name and put into 'wb'
    wb = openpyxl.load_workbook(f.name)
    #close the temporary file
    f.close()
    #return 'wb' to whatever called this function (now go back to the function def import_data below)
    return wb

def getdata(wb, sheet='Sheet1'):
    #Turn a sheet on a workbook into an array
    data = wb.get_sheet_by_name(sheet)
    return [[unicode(cell.value).strip() for cell in row] for row in data.rows]

#This grabs the data from the URL of the spreadsheet
def import_data(linkurl):
    #create variable 'wb' containing the results of running 'getworkbook' on 'linkurl' (the spreadsheet) - see def getworkbook above
    wb = getworkbook(linkurl)
    sheet_names = wb.get_sheet_names() 
    if not sheet_names:
        return
    # assume there's only 1 sheet, and get first one
    rows = getdata(wb, sheet_names[0])
    pprint.pprint(rows)

    expenses = {}
    headings = []
    record = {}
    for num in range(0,18):
#This grabs the headings in row 6 - but the second spreadsheet has these on row 7, breaking the scraper, so the elif line grabs those
        if rows[5][0] == "Name":
            print "rows[5][num]", rows[5][num]
            #various slashes, brackets and full stops caused problems as keys, so this line replaces them all
            cleanedheading = rows[5][num].replace("/","_").replace("No.","Number").replace("(","_").replace(")","_").replace("&","and")
            print "cleaned heading:", cleanedheading 
            headings.append(cleanedheading)
            print "headings:", headings
        elif rows[6][0] == "Name":
            print "rows[6][num]", rows[6][num]
            #various slashes, brackets and full stops caused problems as keys, so this line replaces them all
            cleanedheading = rows[6][num].replace("/","_").replace("No.","Number").replace("(","_").replace(")","_").replace("&","and")
            print "cleaned heading:", cleanedheading 
            headings.append(cleanedheading)
            print "headings:", headings
        else:
            print "NEITHER ROW?", linkurl
        
    for i in range(7,805):
        print rows[i][0]
        expenses['linkurl'] = linkurl
        for num in range(0,18):
            expenses[headings[num]] = rows[i][num]
        #this caused problems because the heading 'Name' has a capital 'n' which I overlooked
        scraperwiki.sqlite.save(['Name', 'linkurl'], expenses, 'expenses')
import scraperwiki
import lxml.html

#This function finds links to all XLSX files on the page passed to it (URL)
def grabexcellinks(URL):
    #Use Scraperwiki's scrape function on 'URL', put results in new variable 'html'
    html = scraperwiki.scrape(URL)
    #Use lxml.html's fromstring function on 'html', put results in new variable 'root'
    root = lxml.html.fromstring(html)
    #use cssselect method on 'root' to grab all <a> tags within a <li> tag within <div class="rte"> - and put in a new list variable 'links'
    links = root.cssselect('div.rte li a')
    #for each item in that list variable, put it in the variable 'link'
    for link in links:
        #and print the text_content of that (after the string "link text:")
        print "link text:", link.text_content()
        #use the attrib.get method on 'link' to grab the href= attribute of the HTML, and put in new 'linkurl' variable
        linkurl = link.attrib.get('href')
        #print it
        print "link type:", linkurl[-4:]
        #if the last 4 characters of linkurl are "xlsx", then:
        if linkurl[-4:] == "xlsx":
            print "SCRAPE IT!"
            #...run the function 'import_data' on the result of adding 'http://www.parliament.uk/' to that URL (see def import_data above)
            import_data('http://www.parliament.uk/'+linkurl)

#Loop through the list 'URLS':
for URL in URLS:
    #and run the function 'grabexcellinks' on each one (see def grabexcellinks above)
    grabexcellinks(URL)

import scraperwiki
import requests
import openpyxl
import tempfile
import time
import pprint

#BASEURL = 'http://www.parliament.uk/'
#URL = 'http://www.parliament.uk/business/lords/whos-in-the-house-of-lords/house-of-lords-expenses/'
#set a variable for the spreadsheet location
#XLS = 'http://www.parliament.uk/documents/lords-finance-office/2012-13/allowances-expenses-2012-13-month-10-january-2013.xlsx'

#This is a list of the 3 URLs we want to scrape. The line at the bottom of this scraper will loop through those.
URLS = ['http://www.parliament.uk/business/lords/whos-in-the-house-of-lords/house-of-lords-expenses/','http://www.parliament.uk/mps-lords-and-offices/members-allowances/house-of-lords/holallowances/hol-expenses04/201112/','http://www.parliament.uk/mps-lords-and-offices/members-allowances/house-of-lords/holallowances/hol-expenses04/201011/']

#create a new function called 'getworkbook' which will take 1 ingredient which it renames 'url'
def getworkbook(url):
    # Loads the xlsx file from 'url' using the 'requests' libary's 'get' and 'content' - see http://docs.python-requests.org/en/latest/user/quickstart/#make-a-request
    raw = requests.get(url, verify=False).content
    #uses the 'tempfile' library and 'NamedTemporaryFile' to store temporary information in a new variable 'f' - see http://docs.python.org/2/library/tempfile.html for more on that library
    f = tempfile.NamedTemporaryFile('wb')
    #uses '.write' to add 'raw' to that
    f.write(raw)
    #uses '.seek' to start from '0' - see http://www.tutorialspoint.com/python/file_seek.htm
    f.seek(0)
    #now use the library openpyxl's .load_workbook function to grab f.name and put into 'wb'
    wb = openpyxl.load_workbook(f.name)
    #close the temporary file
    f.close()
    #return 'wb' to whatever called this function (now go back to the function def import_data below)
    return wb

def getdata(wb, sheet='Sheet1'):
    #Turn a sheet on a workbook into an array
    data = wb.get_sheet_by_name(sheet)
    return [[unicode(cell.value).strip() for cell in row] for row in data.rows]

#This grabs the data from the URL of the spreadsheet
def import_data(linkurl):
    #create variable 'wb' containing the results of running 'getworkbook' on 'linkurl' (the spreadsheet) - see def getworkbook above
    wb = getworkbook(linkurl)
    sheet_names = wb.get_sheet_names() 
    if not sheet_names:
        return
    # assume there's only 1 sheet, and get first one
    rows = getdata(wb, sheet_names[0])
    pprint.pprint(rows)

    expenses = {}
    headings = []
    record = {}
    for num in range(0,18):
#This grabs the headings in row 6 - but the second spreadsheet has these on row 7, breaking the scraper, so the elif line grabs those
        if rows[5][0] == "Name":
            print "rows[5][num]", rows[5][num]
            #various slashes, brackets and full stops caused problems as keys, so this line replaces them all
            cleanedheading = rows[5][num].replace("/","_").replace("No.","Number").replace("(","_").replace(")","_").replace("&","and")
            print "cleaned heading:", cleanedheading 
            headings.append(cleanedheading)
            print "headings:", headings
        elif rows[6][0] == "Name":
            print "rows[6][num]", rows[6][num]
            #various slashes, brackets and full stops caused problems as keys, so this line replaces them all
            cleanedheading = rows[6][num].replace("/","_").replace("No.","Number").replace("(","_").replace(")","_").replace("&","and")
            print "cleaned heading:", cleanedheading 
            headings.append(cleanedheading)
            print "headings:", headings
        else:
            print "NEITHER ROW?", linkurl
        
    for i in range(7,805):
        print rows[i][0]
        expenses['linkurl'] = linkurl
        for num in range(0,18):
            expenses[headings[num]] = rows[i][num]
        #this caused problems because the heading 'Name' has a capital 'n' which I overlooked
        scraperwiki.sqlite.save(['Name', 'linkurl'], expenses, 'expenses')
import scraperwiki
import lxml.html

#This function finds links to all XLSX files on the page passed to it (URL)
def grabexcellinks(URL):
    #Use Scraperwiki's scrape function on 'URL', put results in new variable 'html'
    html = scraperwiki.scrape(URL)
    #Use lxml.html's fromstring function on 'html', put results in new variable 'root'
    root = lxml.html.fromstring(html)
    #use cssselect method on 'root' to grab all <a> tags within a <li> tag within <div class="rte"> - and put in a new list variable 'links'
    links = root.cssselect('div.rte li a')
    #for each item in that list variable, put it in the variable 'link'
    for link in links:
        #and print the text_content of that (after the string "link text:")
        print "link text:", link.text_content()
        #use the attrib.get method on 'link' to grab the href= attribute of the HTML, and put in new 'linkurl' variable
        linkurl = link.attrib.get('href')
        #print it
        print "link type:", linkurl[-4:]
        #if the last 4 characters of linkurl are "xlsx", then:
        if linkurl[-4:] == "xlsx":
            print "SCRAPE IT!"
            #...run the function 'import_data' on the result of adding 'http://www.parliament.uk/' to that URL (see def import_data above)
            import_data('http://www.parliament.uk/'+linkurl)

#Loop through the list 'URLS':
for URL in URLS:
    #and run the function 'grabexcellinks' on each one (see def grabexcellinks above)
    grabexcellinks(URL)

