import scraperwiki
import lxml.html
import requests
import openpyxl
import tempfile
import time
import pprint

BASEURL = 'http://www.parliament.uk/'
URL = 'http://www.parliament.uk/business/lords/whos-in-the-house-of-lords/house-of-lords-expenses/'
#set a variable for the spreadsheet location
XLS = 'http://www.parliament.uk/documents/lords-finance-office/2012-13/allowances-expenses-2012-13-month-10-january-2013.xlsx'

def getworkbook(url):
    # Loads an xlsx file from the internet
    raw = requests.get(url, verify=False).content
    f = tempfile.NamedTemporaryFile('wb')
    f.write(raw)
    f.seek(0)
    wb = openpyxl.load_workbook(f.name)
    f.close()
    return wb

def getdata(wb, sheet='Sheet1'):
    #Turn a sheet on a workbook into an array
    data = wb.get_sheet_by_name(sheet)
    if data:
        return [[unicode(cell.value).strip() for cell in row] for row in data.rows]

def import_data(linkurl):
    wb = getworkbook(linkurl)
    
    #rows = getdata(wb, 'Sheet1')
    sheet_names = wb.get_sheet_names()
    if not sheet_names:
        return
    rows = getdata(wb, sheet_names[0])
    pprint.pprint(rows)
    if not rows:
        return

    expenses = {}
    headings = []
    record = {}
    for num in range(0,18):
#PROBLEM! This grabs the headings in row 6 - but the second spreadsheet has these on row 7, breaking the scraper
        if (len(rows) > 5) and rows[5][0] == "Name":
            print "rows[5][num]", rows[5][num]
            #various slashes, brackets and full stops caused problems as keys, so this line replaces them all
            cleanedheading = rows[5][num].replace("/","_").replace("No.","Number").replace("(","_").replace(")","_").replace("&","and")
            print "cleaned heading:", cleanedheading 
            headings.append(cleanedheading)
            print "headings:", headings
        elif (len(rows) > 6) and rows[6][0] == "Name":
            print "rows[6][num]", rows[6][num]
            #various slashes, brackets and full stops caused problems as keys, so this line replaces them all
            cleanedheading = rows[6][num].replace("/","_").replace("No.","Number").replace("(","_").replace(")","_").replace("&","and")
            print "cleaned heading:", cleanedheading 
            headings.append(cleanedheading)
            print "headings:", headings
        else:
            print "NEITHER ROW?", linkurl



def grabexcellinks(URL):
    #Use Scraperwiki's scrape function on 'URL', put results in new variable 'html'
    html = scraperwiki.scrape(URL)
    #and show it us
    print html
    #Use lxml.html's fromstring function on 'html', put results in new variable 'root'
    root = lxml.html.fromstring(html)
    #use cssselect method on 'root' to grab all <a> tags within a <li> tag - and put in a new list variable 'links'
    links = root.cssselect('div.rte li a')
    #for each item in that list variable, from the [8:23], put it in the variable 'link'
#Works on links up to August (index 11), so running from other links. August (11), July (13), June (15) all break. May (17), April (19) OK
    for link in links[8:23]:
        #and print the text_content of that (after the string "link text:")
        print "link text:", link.text_content()
        #use the attrib.get method on 'link' to grab the href= attribute of the HTML, and put in new 'linkurl' variable
        linkurl = link.attrib.get('href')
        #print it
        print "link type:", linkurl[-4:]
        #run the function scrapesheets, using that variable as the parameter
        if linkurl[-4:] == "xlsx":
            print "SCRAPE IT!"
            import_data('http://www.parliament.uk'+linkurl)
#THIS BIT GRABS THE FINAL WORKING XLSX FILES BUT IGNORE FOR NOW
    for link in links[16:20]:
        #and print the text_content of that (after the string "link text:")
        print "link text:", link.text_content()
        #use the attrib.get method on 'link' to grab the href= attribute of the HTML, and put in new 'linkurl' variable
        linkurl = link.attrib.get('href')
        #print it
        print "link type:", linkurl[-4:]
        #run the function scrapesheets, using that variable as the parameter
        if linkurl[-4:] == "xlsx":
            print "SCRAPE IT!"
            import_data('http://www.parliament.uk'+linkurl)

grabexcellinks('http://www.parliament.uk/business/lords/whos-in-the-house-of-lords/house-of-lords-expenses/')

import scraperwiki
import lxml.html
import requests
import openpyxl
import tempfile
import time
import pprint

BASEURL = 'http://www.parliament.uk/'
URL = 'http://www.parliament.uk/business/lords/whos-in-the-house-of-lords/house-of-lords-expenses/'
#set a variable for the spreadsheet location
XLS = 'http://www.parliament.uk/documents/lords-finance-office/2012-13/allowances-expenses-2012-13-month-10-january-2013.xlsx'

def getworkbook(url):
    # Loads an xlsx file from the internet
    raw = requests.get(url, verify=False).content
    f = tempfile.NamedTemporaryFile('wb')
    f.write(raw)
    f.seek(0)
    wb = openpyxl.load_workbook(f.name)
    f.close()
    return wb

def getdata(wb, sheet='Sheet1'):
    #Turn a sheet on a workbook into an array
    data = wb.get_sheet_by_name(sheet)
    if data:
        return [[unicode(cell.value).strip() for cell in row] for row in data.rows]

def import_data(linkurl):
    wb = getworkbook(linkurl)
    
    #rows = getdata(wb, 'Sheet1')
    sheet_names = wb.get_sheet_names()
    if not sheet_names:
        return
    rows = getdata(wb, sheet_names[0])
    pprint.pprint(rows)
    if not rows:
        return

    expenses = {}
    headings = []
    record = {}
    for num in range(0,18):
#PROBLEM! This grabs the headings in row 6 - but the second spreadsheet has these on row 7, breaking the scraper
        if (len(rows) > 5) and rows[5][0] == "Name":
            print "rows[5][num]", rows[5][num]
            #various slashes, brackets and full stops caused problems as keys, so this line replaces them all
            cleanedheading = rows[5][num].replace("/","_").replace("No.","Number").replace("(","_").replace(")","_").replace("&","and")
            print "cleaned heading:", cleanedheading 
            headings.append(cleanedheading)
            print "headings:", headings
        elif (len(rows) > 6) and rows[6][0] == "Name":
            print "rows[6][num]", rows[6][num]
            #various slashes, brackets and full stops caused problems as keys, so this line replaces them all
            cleanedheading = rows[6][num].replace("/","_").replace("No.","Number").replace("(","_").replace(")","_").replace("&","and")
            print "cleaned heading:", cleanedheading 
            headings.append(cleanedheading)
            print "headings:", headings
        else:
            print "NEITHER ROW?", linkurl



def grabexcellinks(URL):
    #Use Scraperwiki's scrape function on 'URL', put results in new variable 'html'
    html = scraperwiki.scrape(URL)
    #and show it us
    print html
    #Use lxml.html's fromstring function on 'html', put results in new variable 'root'
    root = lxml.html.fromstring(html)
    #use cssselect method on 'root' to grab all <a> tags within a <li> tag - and put in a new list variable 'links'
    links = root.cssselect('div.rte li a')
    #for each item in that list variable, from the [8:23], put it in the variable 'link'
#Works on links up to August (index 11), so running from other links. August (11), July (13), June (15) all break. May (17), April (19) OK
    for link in links[8:23]:
        #and print the text_content of that (after the string "link text:")
        print "link text:", link.text_content()
        #use the attrib.get method on 'link' to grab the href= attribute of the HTML, and put in new 'linkurl' variable
        linkurl = link.attrib.get('href')
        #print it
        print "link type:", linkurl[-4:]
        #run the function scrapesheets, using that variable as the parameter
        if linkurl[-4:] == "xlsx":
            print "SCRAPE IT!"
            import_data('http://www.parliament.uk'+linkurl)
#THIS BIT GRABS THE FINAL WORKING XLSX FILES BUT IGNORE FOR NOW
    for link in links[16:20]:
        #and print the text_content of that (after the string "link text:")
        print "link text:", link.text_content()
        #use the attrib.get method on 'link' to grab the href= attribute of the HTML, and put in new 'linkurl' variable
        linkurl = link.attrib.get('href')
        #print it
        print "link type:", linkurl[-4:]
        #run the function scrapesheets, using that variable as the parameter
        if linkurl[-4:] == "xlsx":
            print "SCRAPE IT!"
            import_data('http://www.parliament.uk'+linkurl)

grabexcellinks('http://www.parliament.uk/business/lords/whos-in-the-house-of-lords/house-of-lords-expenses/')

import scraperwiki
import lxml.html
import requests
import openpyxl
import tempfile
import time
import pprint

BASEURL = 'http://www.parliament.uk/'
URL = 'http://www.parliament.uk/business/lords/whos-in-the-house-of-lords/house-of-lords-expenses/'
#set a variable for the spreadsheet location
XLS = 'http://www.parliament.uk/documents/lords-finance-office/2012-13/allowances-expenses-2012-13-month-10-january-2013.xlsx'

def getworkbook(url):
    # Loads an xlsx file from the internet
    raw = requests.get(url, verify=False).content
    f = tempfile.NamedTemporaryFile('wb')
    f.write(raw)
    f.seek(0)
    wb = openpyxl.load_workbook(f.name)
    f.close()
    return wb

def getdata(wb, sheet='Sheet1'):
    #Turn a sheet on a workbook into an array
    data = wb.get_sheet_by_name(sheet)
    if data:
        return [[unicode(cell.value).strip() for cell in row] for row in data.rows]

def import_data(linkurl):
    wb = getworkbook(linkurl)
    
    #rows = getdata(wb, 'Sheet1')
    sheet_names = wb.get_sheet_names()
    if not sheet_names:
        return
    rows = getdata(wb, sheet_names[0])
    pprint.pprint(rows)
    if not rows:
        return

    expenses = {}
    headings = []
    record = {}
    for num in range(0,18):
#PROBLEM! This grabs the headings in row 6 - but the second spreadsheet has these on row 7, breaking the scraper
        if (len(rows) > 5) and rows[5][0] == "Name":
            print "rows[5][num]", rows[5][num]
            #various slashes, brackets and full stops caused problems as keys, so this line replaces them all
            cleanedheading = rows[5][num].replace("/","_").replace("No.","Number").replace("(","_").replace(")","_").replace("&","and")
            print "cleaned heading:", cleanedheading 
            headings.append(cleanedheading)
            print "headings:", headings
        elif (len(rows) > 6) and rows[6][0] == "Name":
            print "rows[6][num]", rows[6][num]
            #various slashes, brackets and full stops caused problems as keys, so this line replaces them all
            cleanedheading = rows[6][num].replace("/","_").replace("No.","Number").replace("(","_").replace(")","_").replace("&","and")
            print "cleaned heading:", cleanedheading 
            headings.append(cleanedheading)
            print "headings:", headings
        else:
            print "NEITHER ROW?", linkurl



def grabexcellinks(URL):
    #Use Scraperwiki's scrape function on 'URL', put results in new variable 'html'
    html = scraperwiki.scrape(URL)
    #and show it us
    print html
    #Use lxml.html's fromstring function on 'html', put results in new variable 'root'
    root = lxml.html.fromstring(html)
    #use cssselect method on 'root' to grab all <a> tags within a <li> tag - and put in a new list variable 'links'
    links = root.cssselect('div.rte li a')
    #for each item in that list variable, from the [8:23], put it in the variable 'link'
#Works on links up to August (index 11), so running from other links. August (11), July (13), June (15) all break. May (17), April (19) OK
    for link in links[8:23]:
        #and print the text_content of that (after the string "link text:")
        print "link text:", link.text_content()
        #use the attrib.get method on 'link' to grab the href= attribute of the HTML, and put in new 'linkurl' variable
        linkurl = link.attrib.get('href')
        #print it
        print "link type:", linkurl[-4:]
        #run the function scrapesheets, using that variable as the parameter
        if linkurl[-4:] == "xlsx":
            print "SCRAPE IT!"
            import_data('http://www.parliament.uk'+linkurl)
#THIS BIT GRABS THE FINAL WORKING XLSX FILES BUT IGNORE FOR NOW
    for link in links[16:20]:
        #and print the text_content of that (after the string "link text:")
        print "link text:", link.text_content()
        #use the attrib.get method on 'link' to grab the href= attribute of the HTML, and put in new 'linkurl' variable
        linkurl = link.attrib.get('href')
        #print it
        print "link type:", linkurl[-4:]
        #run the function scrapesheets, using that variable as the parameter
        if linkurl[-4:] == "xlsx":
            print "SCRAPE IT!"
            import_data('http://www.parliament.uk'+linkurl)

grabexcellinks('http://www.parliament.uk/business/lords/whos-in-the-house-of-lords/house-of-lords-expenses/')

