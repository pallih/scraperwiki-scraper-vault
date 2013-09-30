"""
This is an example of scraping multiple URLs.
"""

import scraperwiki
import re
import xlwt

from openpyxl import Workbook
wb = xlwt.Workbook(encoding="utf-8")

sheet1 = wb.add_sheet("Parking1") 

# The URLs we're going to scrape:

urls = """

http://www.nationalrail.co.uk/stations/abw/details.html
http://www.nationalrail.co.uk/stations/RAU/details.html

""".strip()

urls = urls.splitlines()


# This is a useful helper function that you might want to steal.
# It cleans up the data a bit.

#def gettext(html):
    #"""Return the text within html, removing any HTML tags it contained."""
    #cleaned = re.sub('<.*?>', '', html)  # remove tags
    #cleaned = ' '.join(cleaned.split())  # collapse whitespace
    #return cleaned


for url in urls:
    #print "Scraping", url
    page = scraperwiki.scrape(url)
    if page is not None:
        # Store all the h2 headings (matching "<h2>...........</h2>")
        # headings = re.findall("<br /><h3>(.*?)</h3>", page, re.DOTALL)
        headings = re.findall("<p><strong>Charges:</strong>(.*?)</p>", page, re.DOTALL)

        # re.DOTALL makes it work across multiple lines as well.

        # Just keep the heading text
        # (Try commenting this out)
        # headings = [gettext(heading) for heading in headings]

        data = {'url': url, 'headings': headings}
        scraperwiki.sqlite.save(['url'], data)   # each entry is identified by its url

sheet1.write(0, 0, "This is the First Cell of the First Sheet")

wb.save('Parking.xlsx')

#from openpyxl import load_workbook
#wb2 = load_workbook('test.xlsx')
#print wb2.get_sheet_names()
#['Sheet2', 'New Title', 'Sheet1']

"""
This is an example of scraping multiple URLs.
"""

import scraperwiki
import re
import xlwt

from openpyxl import Workbook
wb = xlwt.Workbook(encoding="utf-8")

sheet1 = wb.add_sheet("Parking1") 

# The URLs we're going to scrape:

urls = """

http://www.nationalrail.co.uk/stations/abw/details.html
http://www.nationalrail.co.uk/stations/RAU/details.html

""".strip()

urls = urls.splitlines()


# This is a useful helper function that you might want to steal.
# It cleans up the data a bit.

#def gettext(html):
    #"""Return the text within html, removing any HTML tags it contained."""
    #cleaned = re.sub('<.*?>', '', html)  # remove tags
    #cleaned = ' '.join(cleaned.split())  # collapse whitespace
    #return cleaned


for url in urls:
    #print "Scraping", url
    page = scraperwiki.scrape(url)
    if page is not None:
        # Store all the h2 headings (matching "<h2>...........</h2>")
        # headings = re.findall("<br /><h3>(.*?)</h3>", page, re.DOTALL)
        headings = re.findall("<p><strong>Charges:</strong>(.*?)</p>", page, re.DOTALL)

        # re.DOTALL makes it work across multiple lines as well.

        # Just keep the heading text
        # (Try commenting this out)
        # headings = [gettext(heading) for heading in headings]

        data = {'url': url, 'headings': headings}
        scraperwiki.sqlite.save(['url'], data)   # each entry is identified by its url

sheet1.write(0, 0, "This is the First Cell of the First Sheet")

wb.save('Parking.xlsx')

#from openpyxl import load_workbook
#wb2 = load_workbook('test.xlsx')
#print wb2.get_sheet_names()
#['Sheet2', 'New Title', 'Sheet1']

