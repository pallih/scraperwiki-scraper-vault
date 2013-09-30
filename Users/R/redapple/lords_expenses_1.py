import scraperwiki
import requests
import openpyxl
import tempfile
import time
import pprint
import scraperwiki
import lxml.html


BASEURL = 'http://www.parliament.uk/'
URL = 'http://www.parliament.uk/business/lords/whos-in-the-house-of-lords/house-of-lords-expenses/'

def getworkbook(url):
    # Loads an xlsx file from the internet
    raw = requests.get(url, verify=False).content
    f = tempfile.NamedTemporaryFile('wb')
    f.write(raw)
    f.seek(0)
    wb = openpyxl.load_workbook(f.name)
    f.close()
    return wb

def getdata(wb, sheet='Sheet1'):
    #Turn a sheet on a workbook into an array
    data = wb.get_sheet_by_name(sheet)
    return [[unicode(cell.value).strip() for cell in row] for row in data.rows]

def import_data(linkurl):
    wb = getworkbook(linkurl)
    sheet_names = wb.get_sheet_names() 
    if not sheet_names:
        return
    
    # assume there's only 1 sheet, and get first one
    rows = getdata(wb, sheet_names[0])
    pprint.pprint(rows)

    expenses = {}
    headings = []
    record = {}
    for num in range(0,18):
        #This grabs the headings in row 6 - but the second spreadsheet has these on row 7, breaking the scraper, so the elif line grabs those
        if rows[5][0] == "Name":
            target_row = rows[5]
        elif rows[6][0] == "Name":
            target_row = rows[6]
        else:
            print "NEITHER ROW?", linkurl
            continue
        print "rows[5/6][num]", target_row[num]
        #various slashes, brackets and full stops caused problems as keys, so this line replaces them all
        cleanedheading = target_row[num].replace("/","_").replace("No.","Number").replace("(","_").replace(")","_").replace("&","and")
        print "cleaned heading:", cleanedheading 
        headings.append(cleanedheading)
        print "headings:", headings
        
    for i in range(7,805):
        print rows[i][0]
        expenses['linkurl'] = linkurl
        for num in range(0,18):
            expenses[headings[num]] = rows[i][num]
        #this caused problems because the heading 'Name' has a capital 'n' which I overlooked
        scraperwiki.sqlite.save(['Name', 'linkurl'], expenses, 'expenses')


def grabexcellinks(URL):

    #Use Scraperwiki's scrape function on 'URL', put results in new variable 'html'
    html = scraperwiki.scrape(URL)

    #and show it us
    print html

    #Use lxml.html's fromstring function on 'html', put results in new variable 'root'
    root = lxml.html.fromstring(html)

    # use cssselect method on 'root' to grab all <a> tags within a <li> tag - and put in a new list variable 'links'
    # here we select link that end with XLSX extension directly
    # http://www.w3.org/TR/css3-selectors/#attribute-substrings
    for link in root.cssselect('div.rte li a[href$=".xlsx"]'):

        #and print the text_content of that (after the string "link text:")
        print "link text:", link.text_content()

        #use the attrib.get method on 'link' to grab the href= attribute of the HTML, and put in new 'linkurl' variable
        linkurl = link.attrib.get('href')

        print "SCRAPE IT!"
        import_data('http://www.parliament.uk/'+linkurl)

grabexcellinks('http://www.parliament.uk/business/lords/whos-in-the-house-of-lords/house-of-lords-expenses/')

import scraperwiki
import requests
import openpyxl
import tempfile
import time
import pprint
import scraperwiki
import lxml.html


BASEURL = 'http://www.parliament.uk/'
URL = 'http://www.parliament.uk/business/lords/whos-in-the-house-of-lords/house-of-lords-expenses/'

def getworkbook(url):
    # Loads an xlsx file from the internet
    raw = requests.get(url, verify=False).content
    f = tempfile.NamedTemporaryFile('wb')
    f.write(raw)
    f.seek(0)
    wb = openpyxl.load_workbook(f.name)
    f.close()
    return wb

def getdata(wb, sheet='Sheet1'):
    #Turn a sheet on a workbook into an array
    data = wb.get_sheet_by_name(sheet)
    return [[unicode(cell.value).strip() for cell in row] for row in data.rows]

def import_data(linkurl):
    wb = getworkbook(linkurl)
    sheet_names = wb.get_sheet_names() 
    if not sheet_names:
        return
    
    # assume there's only 1 sheet, and get first one
    rows = getdata(wb, sheet_names[0])
    pprint.pprint(rows)

    expenses = {}
    headings = []
    record = {}
    for num in range(0,18):
        #This grabs the headings in row 6 - but the second spreadsheet has these on row 7, breaking the scraper, so the elif line grabs those
        if rows[5][0] == "Name":
            target_row = rows[5]
        elif rows[6][0] == "Name":
            target_row = rows[6]
        else:
            print "NEITHER ROW?", linkurl
            continue
        print "rows[5/6][num]", target_row[num]
        #various slashes, brackets and full stops caused problems as keys, so this line replaces them all
        cleanedheading = target_row[num].replace("/","_").replace("No.","Number").replace("(","_").replace(")","_").replace("&","and")
        print "cleaned heading:", cleanedheading 
        headings.append(cleanedheading)
        print "headings:", headings
        
    for i in range(7,805):
        print rows[i][0]
        expenses['linkurl'] = linkurl
        for num in range(0,18):
            expenses[headings[num]] = rows[i][num]
        #this caused problems because the heading 'Name' has a capital 'n' which I overlooked
        scraperwiki.sqlite.save(['Name', 'linkurl'], expenses, 'expenses')


def grabexcellinks(URL):

    #Use Scraperwiki's scrape function on 'URL', put results in new variable 'html'
    html = scraperwiki.scrape(URL)

    #and show it us
    print html

    #Use lxml.html's fromstring function on 'html', put results in new variable 'root'
    root = lxml.html.fromstring(html)

    # use cssselect method on 'root' to grab all <a> tags within a <li> tag - and put in a new list variable 'links'
    # here we select link that end with XLSX extension directly
    # http://www.w3.org/TR/css3-selectors/#attribute-substrings
    for link in root.cssselect('div.rte li a[href$=".xlsx"]'):

        #and print the text_content of that (after the string "link text:")
        print "link text:", link.text_content()

        #use the attrib.get method on 'link' to grab the href= attribute of the HTML, and put in new 'linkurl' variable
        linkurl = link.attrib.get('href')

        print "SCRAPE IT!"
        import_data('http://www.parliament.uk/'+linkurl)

grabexcellinks('http://www.parliament.uk/business/lords/whos-in-the-house-of-lords/house-of-lords-expenses/')

import scraperwiki
import requests
import openpyxl
import tempfile
import time
import pprint
import scraperwiki
import lxml.html


BASEURL = 'http://www.parliament.uk/'
URL = 'http://www.parliament.uk/business/lords/whos-in-the-house-of-lords/house-of-lords-expenses/'

def getworkbook(url):
    # Loads an xlsx file from the internet
    raw = requests.get(url, verify=False).content
    f = tempfile.NamedTemporaryFile('wb')
    f.write(raw)
    f.seek(0)
    wb = openpyxl.load_workbook(f.name)
    f.close()
    return wb

def getdata(wb, sheet='Sheet1'):
    #Turn a sheet on a workbook into an array
    data = wb.get_sheet_by_name(sheet)
    return [[unicode(cell.value).strip() for cell in row] for row in data.rows]

def import_data(linkurl):
    wb = getworkbook(linkurl)
    sheet_names = wb.get_sheet_names() 
    if not sheet_names:
        return
    
    # assume there's only 1 sheet, and get first one
    rows = getdata(wb, sheet_names[0])
    pprint.pprint(rows)

    expenses = {}
    headings = []
    record = {}
    for num in range(0,18):
        #This grabs the headings in row 6 - but the second spreadsheet has these on row 7, breaking the scraper, so the elif line grabs those
        if rows[5][0] == "Name":
            target_row = rows[5]
        elif rows[6][0] == "Name":
            target_row = rows[6]
        else:
            print "NEITHER ROW?", linkurl
            continue
        print "rows[5/6][num]", target_row[num]
        #various slashes, brackets and full stops caused problems as keys, so this line replaces them all
        cleanedheading = target_row[num].replace("/","_").replace("No.","Number").replace("(","_").replace(")","_").replace("&","and")
        print "cleaned heading:", cleanedheading 
        headings.append(cleanedheading)
        print "headings:", headings
        
    for i in range(7,805):
        print rows[i][0]
        expenses['linkurl'] = linkurl
        for num in range(0,18):
            expenses[headings[num]] = rows[i][num]
        #this caused problems because the heading 'Name' has a capital 'n' which I overlooked
        scraperwiki.sqlite.save(['Name', 'linkurl'], expenses, 'expenses')


def grabexcellinks(URL):

    #Use Scraperwiki's scrape function on 'URL', put results in new variable 'html'
    html = scraperwiki.scrape(URL)

    #and show it us
    print html

    #Use lxml.html's fromstring function on 'html', put results in new variable 'root'
    root = lxml.html.fromstring(html)

    # use cssselect method on 'root' to grab all <a> tags within a <li> tag - and put in a new list variable 'links'
    # here we select link that end with XLSX extension directly
    # http://www.w3.org/TR/css3-selectors/#attribute-substrings
    for link in root.cssselect('div.rte li a[href$=".xlsx"]'):

        #and print the text_content of that (after the string "link text:")
        print "link text:", link.text_content()

        #use the attrib.get method on 'link' to grab the href= attribute of the HTML, and put in new 'linkurl' variable
        linkurl = link.attrib.get('href')

        print "SCRAPE IT!"
        import_data('http://www.parliament.uk/'+linkurl)

grabexcellinks('http://www.parliament.uk/business/lords/whos-in-the-house-of-lords/house-of-lords-expenses/')

import scraperwiki
import requests
import openpyxl
import tempfile
import time
import pprint
import scraperwiki
import lxml.html


BASEURL = 'http://www.parliament.uk/'
URL = 'http://www.parliament.uk/business/lords/whos-in-the-house-of-lords/house-of-lords-expenses/'

def getworkbook(url):
    # Loads an xlsx file from the internet
    raw = requests.get(url, verify=False).content
    f = tempfile.NamedTemporaryFile('wb')
    f.write(raw)
    f.seek(0)
    wb = openpyxl.load_workbook(f.name)
    f.close()
    return wb

def getdata(wb, sheet='Sheet1'):
    #Turn a sheet on a workbook into an array
    data = wb.get_sheet_by_name(sheet)
    return [[unicode(cell.value).strip() for cell in row] for row in data.rows]

def import_data(linkurl):
    wb = getworkbook(linkurl)
    sheet_names = wb.get_sheet_names() 
    if not sheet_names:
        return
    
    # assume there's only 1 sheet, and get first one
    rows = getdata(wb, sheet_names[0])
    pprint.pprint(rows)

    expenses = {}
    headings = []
    record = {}
    for num in range(0,18):
        #This grabs the headings in row 6 - but the second spreadsheet has these on row 7, breaking the scraper, so the elif line grabs those
        if rows[5][0] == "Name":
            target_row = rows[5]
        elif rows[6][0] == "Name":
            target_row = rows[6]
        else:
            print "NEITHER ROW?", linkurl
            continue
        print "rows[5/6][num]", target_row[num]
        #various slashes, brackets and full stops caused problems as keys, so this line replaces them all
        cleanedheading = target_row[num].replace("/","_").replace("No.","Number").replace("(","_").replace(")","_").replace("&","and")
        print "cleaned heading:", cleanedheading 
        headings.append(cleanedheading)
        print "headings:", headings
        
    for i in range(7,805):
        print rows[i][0]
        expenses['linkurl'] = linkurl
        for num in range(0,18):
            expenses[headings[num]] = rows[i][num]
        #this caused problems because the heading 'Name' has a capital 'n' which I overlooked
        scraperwiki.sqlite.save(['Name', 'linkurl'], expenses, 'expenses')


def grabexcellinks(URL):

    #Use Scraperwiki's scrape function on 'URL', put results in new variable 'html'
    html = scraperwiki.scrape(URL)

    #and show it us
    print html

    #Use lxml.html's fromstring function on 'html', put results in new variable 'root'
    root = lxml.html.fromstring(html)

    # use cssselect method on 'root' to grab all <a> tags within a <li> tag - and put in a new list variable 'links'
    # here we select link that end with XLSX extension directly
    # http://www.w3.org/TR/css3-selectors/#attribute-substrings
    for link in root.cssselect('div.rte li a[href$=".xlsx"]'):

        #and print the text_content of that (after the string "link text:")
        print "link text:", link.text_content()

        #use the attrib.get method on 'link' to grab the href= attribute of the HTML, and put in new 'linkurl' variable
        linkurl = link.attrib.get('href')

        print "SCRAPE IT!"
        import_data('http://www.parliament.uk/'+linkurl)

grabexcellinks('http://www.parliament.uk/business/lords/whos-in-the-house-of-lords/house-of-lords-expenses/')

