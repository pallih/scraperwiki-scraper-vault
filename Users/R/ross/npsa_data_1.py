import scraperwiki
import lxml.html
import urlparse
import re
from openpyxl import load_workbook

site = 'http://www.nrls.npsa.nhs.uk/patient-safety-data/'
cols = {
    1 : "SHA name",
    2 : "Trust name",
    3 : "Alert title",
    4 : "Alert reference",
    5 : "Issue date",
    6 : "Completion deadline date",
    7 : "Current Status"
}

def get_month_year( s ):
    m = re.match(r'.*\xa0(\w+)\xa0(\d+)', s)
    if m:
        return m.groups(0)[0],m.groups(0)[1]
    return "", ""

def xls_list():
    result = []
    html = scraperwiki.scrape(site)
    page = lxml.html.fromstring( html )
    links = page.cssselect('a.oLinkAsset') + page.cssselect('a.oLinkAssetXls')
    for link in links:
        name = link.text_content()
        href = urlparse.urljoin(site, link.attrib.get('href'))
        m,y = get_month_year( name )
        result.append( ( name, m, y, href, ) )
    return [result[0]]


for name,m,y,l in xls_list():
    print 'Fetching', name
    with open('/tmp/t.xls', 'wb') as f:
        f.write( scraperwiki.scrape(l) )
    print 'Loading xls'
    wb = load_workbook(filename = '/tmp/t.xls')
    sheet = wb.get_active_sheet()
    #print sheet.cell('A5').value 
    
    print 'Processing data'
    for row in range(5, sheet.get_highest_row() + 6 ):
        drow = {}
        for c in range(1, 8 ):
            drow[ cols[c] ] = sheet.cell(row=row, column=c ).value
        drow['year'] = y
        drow['month'] = m
        scraperwiki.sqlite.save( ['Alert title', 'SHA name' ], drow )


