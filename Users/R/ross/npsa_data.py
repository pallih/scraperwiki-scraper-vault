import scraperwiki
import lxml.html
import urlparse
import re
from openpyxl import load_workbook

site = 'http://www.nrls.npsa.nhs.uk/patient-safety-data/'
def get_month_year( s ):
    m = re.match(r'.*\xa0(\w+)\xa0(\d+)', s)
    return (m.groups(0)[0],m.groups(0)[1],) if m else ("","",)

def xls_list():
    result = []
    html = scraperwiki.scrape(site)
    page = lxml.html.fromstring( html )
    links = page.cssselect('a.oLinkAsset') + page.cssselect('a.oLinkAssetXls')
    for link in links:
        name = link.text_content()
        href = urlparse.urljoin(site, link.attrib.get('href'))
        m,y = get_month_year( name )
        result.append( ( name, m, y, href, ) )
    return result


for name,m,y,l in xls_list():
    print 'Processing new file', name
    with open('/tmp/t.xls', 'wb') as f:
        f.write( scraperwiki.scrape(l) )

    wb = load_workbook(filename = '/tmp/t.xls')
    sheet = wb.get_active_sheet()
    written = 0
    headers = []
    for row in sheet.rows:
        if row[3].value is None:
            print 'Skipping'
            continue

        if not headers:
            headers = [ x.value for x in row ]
            continue

        drow = { 'Year': y, 'Month': m }
        for idx, cell in enumerate(row):
            drow[ headers[idx] ] = cell.value

        scraperwiki.sqlite.save( [u'Alert title', u'SHA name' ], drow, verbose=0 )
        written = written + 1
        if written % 100 == 0:
            print written
import scraperwiki
import lxml.html
import urlparse
import re
from openpyxl import load_workbook

site = 'http://www.nrls.npsa.nhs.uk/patient-safety-data/'
def get_month_year( s ):
    m = re.match(r'.*\xa0(\w+)\xa0(\d+)', s)
    return (m.groups(0)[0],m.groups(0)[1],) if m else ("","",)

def xls_list():
    result = []
    html = scraperwiki.scrape(site)
    page = lxml.html.fromstring( html )
    links = page.cssselect('a.oLinkAsset') + page.cssselect('a.oLinkAssetXls')
    for link in links:
        name = link.text_content()
        href = urlparse.urljoin(site, link.attrib.get('href'))
        m,y = get_month_year( name )
        result.append( ( name, m, y, href, ) )
    return result


for name,m,y,l in xls_list():
    print 'Processing new file', name
    with open('/tmp/t.xls', 'wb') as f:
        f.write( scraperwiki.scrape(l) )

    wb = load_workbook(filename = '/tmp/t.xls')
    sheet = wb.get_active_sheet()
    written = 0
    headers = []
    for row in sheet.rows:
        if row[3].value is None:
            print 'Skipping'
            continue

        if not headers:
            headers = [ x.value for x in row ]
            continue

        drow = { 'Year': y, 'Month': m }
        for idx, cell in enumerate(row):
            drow[ headers[idx] ] = cell.value

        scraperwiki.sqlite.save( [u'Alert title', u'SHA name' ], drow, verbose=0 )
        written = written + 1
        if written % 100 == 0:
            print written
import scraperwiki
import lxml.html
import urlparse
import re
from openpyxl import load_workbook

site = 'http://www.nrls.npsa.nhs.uk/patient-safety-data/'
def get_month_year( s ):
    m = re.match(r'.*\xa0(\w+)\xa0(\d+)', s)
    return (m.groups(0)[0],m.groups(0)[1],) if m else ("","",)

def xls_list():
    result = []
    html = scraperwiki.scrape(site)
    page = lxml.html.fromstring( html )
    links = page.cssselect('a.oLinkAsset') + page.cssselect('a.oLinkAssetXls')
    for link in links:
        name = link.text_content()
        href = urlparse.urljoin(site, link.attrib.get('href'))
        m,y = get_month_year( name )
        result.append( ( name, m, y, href, ) )
    return result


for name,m,y,l in xls_list():
    print 'Processing new file', name
    with open('/tmp/t.xls', 'wb') as f:
        f.write( scraperwiki.scrape(l) )

    wb = load_workbook(filename = '/tmp/t.xls')
    sheet = wb.get_active_sheet()
    written = 0
    headers = []
    for row in sheet.rows:
        if row[3].value is None:
            print 'Skipping'
            continue

        if not headers:
            headers = [ x.value for x in row ]
            continue

        drow = { 'Year': y, 'Month': m }
        for idx, cell in enumerate(row):
            drow[ headers[idx] ] = cell.value

        scraperwiki.sqlite.save( [u'Alert title', u'SHA name' ], drow, verbose=0 )
        written = written + 1
        if written % 100 == 0:
            print written
