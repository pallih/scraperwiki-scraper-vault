import scraperwiki
from openpyxl.reader.excel import load_workbook


blob = scraperwiki.scrape("http://www.ic.nhs.uk/catalogue/PUB08354/hosp-pres-eng-2011-Table4.xlsx")
open('/tmp/data.xlsx', 'wb').write(blob)
wb = load_workbook(filename = '/tmp/data.xlsx')

sheet = wb.get_sheet_by_name('Table4')
for row in range(2,153):
    name = sheet.cell('A' + str(row)).value
    primary_care = float(sheet.cell('B' + str(row)).value)
    fp10hp   = float(sheet.cell('D' + str(row)).value) 
    hospital = float(sheet.cell('F' + str(row)).value)
    total =  float(sheet.cell('H' + str(row)).value)

    scraperwiki.sqlite.save(['name'], {'name':name, 'primary_care':primary_care, 
                                       'fp10hp': fp10hp, 'hospital': hospital,
                                        'total': total}, table_name='spending')import scraperwiki
from openpyxl.reader.excel import load_workbook


blob = scraperwiki.scrape("http://www.ic.nhs.uk/catalogue/PUB08354/hosp-pres-eng-2011-Table4.xlsx")
open('/tmp/data.xlsx', 'wb').write(blob)
wb = load_workbook(filename = '/tmp/data.xlsx')

sheet = wb.get_sheet_by_name('Table4')
for row in range(2,153):
    name = sheet.cell('A' + str(row)).value
    primary_care = float(sheet.cell('B' + str(row)).value)
    fp10hp   = float(sheet.cell('D' + str(row)).value) 
    hospital = float(sheet.cell('F' + str(row)).value)
    total =  float(sheet.cell('H' + str(row)).value)

    scraperwiki.sqlite.save(['name'], {'name':name, 'primary_care':primary_care, 
                                       'fp10hp': fp10hp, 'hospital': hospital,
                                        'total': total}, table_name='spending')import scraperwiki
from openpyxl.reader.excel import load_workbook


blob = scraperwiki.scrape("http://www.ic.nhs.uk/catalogue/PUB08354/hosp-pres-eng-2011-Table4.xlsx")
open('/tmp/data.xlsx', 'wb').write(blob)
wb = load_workbook(filename = '/tmp/data.xlsx')

sheet = wb.get_sheet_by_name('Table4')
for row in range(2,153):
    name = sheet.cell('A' + str(row)).value
    primary_care = float(sheet.cell('B' + str(row)).value)
    fp10hp   = float(sheet.cell('D' + str(row)).value) 
    hospital = float(sheet.cell('F' + str(row)).value)
    total =  float(sheet.cell('H' + str(row)).value)

    scraperwiki.sqlite.save(['name'], {'name':name, 'primary_care':primary_care, 
                                       'fp10hp': fp10hp, 'hospital': hospital,
                                        'total': total}, table_name='spending')import scraperwiki
from openpyxl.reader.excel import load_workbook


blob = scraperwiki.scrape("http://www.ic.nhs.uk/catalogue/PUB08354/hosp-pres-eng-2011-Table4.xlsx")
open('/tmp/data.xlsx', 'wb').write(blob)
wb = load_workbook(filename = '/tmp/data.xlsx')

sheet = wb.get_sheet_by_name('Table4')
for row in range(2,153):
    name = sheet.cell('A' + str(row)).value
    primary_care = float(sheet.cell('B' + str(row)).value)
    fp10hp   = float(sheet.cell('D' + str(row)).value) 
    hospital = float(sheet.cell('F' + str(row)).value)
    total =  float(sheet.cell('H' + str(row)).value)

    scraperwiki.sqlite.save(['name'], {'name':name, 'primary_care':primary_care, 
                                       'fp10hp': fp10hp, 'hospital': hospital,
                                        'total': total}, table_name='spending')